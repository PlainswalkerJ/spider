import re
import time
from lxml import etree
from selenium import webdriver
import openpyxl

def url_get(url):
    print("########################")
    print("Spider operational!")
    driver = webdriver.PhantomJS("C:/phantomjs/bin/phantomjs.exe")
    driver.get(url)
    time.sleep(10)
    content = driver.page_source.encode('utf-8')
    time.sleep(10)
    print('Just finished!')
    return content

def save_data(data,filename):
    print('save %s.'%filename)
    with open(filename,'wb') as f:
        f.write(data)
    print('Data has been stored!')
    print("########################")

def do_spider(url):
    html=url_get(url)
    xml = etree.HTML(html)
    class_specs=url[70:].split('&')
    bossname=xml.xpath('//*[@id="filter-boss-text"]')
    difficultyoption=xml.xpath('//*[@id="filter-difficulty-text"]')
    classes=xml.xpath('//*[@id="class-'+class_specs[0]+'"]')
    specs=xml.xpath('//*[@id="class-'+class_specs[0]+'-spec-'+class_specs[1][5:]+'"]')
    spec=specs[0].xpath('string(.)')
    boss_name=bossname[0].xpath('string(.)')
    difficulty=difficultyoption[0].xpath('string(.)')
    clas=classes[0].xpath('string(.)')
    filename=boss_name+'_'+difficulty+'_'+spec+'_'+clas+'.html'
    #################################################################
    wb = openpyxl.Workbook()
    ws=wb.active
    ws['A1'] = '玩家ID'
    ws['B1'] = 'DPS'
    ws['C1'] = '橙装1'
    ws['D1'] = '橙装2'
    for players in range(1,101):
        index=players
        idpath='//*[@id="row-'+url[46:50]+'-'+str(index)+'"]/td[2]/div/div[1]/a[2]'
        playerID = xml.xpath(idpath)
        ID=playerID[0].xpath('string(.)')
        dpspath='//*[@id="row-'+url[46:50]+'-'+str(index)+'"]/td[4]'
        playerDPS=xml.xpath(dpspath)
        DPS=playerDPS[0].xpath('string(.)')

        playerGear=xml.xpath('//*[@id="row-'+url[46:50]+'-'+str(players)+'"]/script[2]/text()')
        ws['A'+str(players+1)]=str(ID)
        ws['B' + str(players + 1)] = str(re.findall('(?<=\t).+(?=\t)',DPS)[0])
        gears=str(playerGear[0])
        GEAR=gears.split(';')
        count=0
        for line in range(0, len(GEAR)):
            if re.findall(r'(?<=name: ").+?(?=", quality: "legendary")',GEAR[line]) != [] and count==0 :
                ws['C' + str(players + 1)] = str(re.findall(r'(?<=name: ").+?(?=", quality: "legendary")',GEAR[line])[0])
                count=1
            elif re.findall(r'(?<=name: ").+?(?=", quality: "legendary")',GEAR[line]) != [] and count==1 :
                ws['D' + str(players + 1)] = str(re.findall(r'(?<=name: ").+?(?=", quality: "legendary")', GEAR[line])[0])
                count=0

    wb.save(boss_name+'_'+difficulty+'_'+spec+'_'+clas+'.xlsx')
    print('Workbook complete!')
    save_data(html,filename)

if __name__ == '__main__':
    classes=['DeathKnight','DemonHunter','Druid','Hunter','Mage','Monk','Paladin','Priest','Rogue','Shaman','Warlock','Warrior']
    spec=[['Frost','Unholy'],['Havoc'],['Balance','Feral'],['BeastMastery','Marksmanship','Survival']
         ,['Arcane','Fire','Frost'],['Windwalker'],['Retribution'],['Shadow'],['Assassination','Outlaw','Subtlety']
          ,['Elemental','Enhancement'],['Affliction','Demonology','Destruction'],['Arms','Fury']]
    boss=['2032','2048','2036','2037','2050','2054','2052','2038','2051']
    t=0
    for i in range(0,9):
         for k in range(0,12):
            for j in range(0,len(spec[k])):
                url = 'https://www.warcraftlogs.com/rankings/13#boss=' + boss[i] + '&difficulty=' + '5' + '&class=' + classes[k] + '&spec=' + spec[k][j]
                do_spider(url)
                time.sleep(5)



